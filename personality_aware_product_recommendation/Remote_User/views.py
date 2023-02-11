from django.db.models import Count
from django.db.models import Q
from django.shortcuts import render, redirect, get_object_or_404
import datetime
import openpyxl
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import re
from sklearn.ensemble import VotingClassifier

import warnings
warnings.filterwarnings("ignore")
plt.style.use('ggplot')
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.metrics import accuracy_score, confusion_matrix, classification_report
from sklearn.metrics import accuracy_score
from sklearn.metrics import f1_score




# Create your views here.
from Remote_User.models import ClientRegister_Model,Product_Details,Recommend_Prediction,detection_ratio,detection_accuracy

def login(request):


    if request.method == "POST" and 'submit1' in request.POST:

        username = request.POST.get('username')
        password = request.POST.get('password')
        try:
            enter = ClientRegister_Model.objects.get(username=username,password=password)
            request.session["userid"] = enter.id

            return redirect('Add_DataSet_Details')
        except:
            pass

    return render(request,'RUser/login.html')

def Add_DataSet_Details(request):
    if "GET" == request.method:
        return render(request, 'RUser/Add_DataSet_Details.html', {})
    else:
        excel_file = request.FILES["excel_file"]
        # you may put validations here to check extension or file size
        wb = openpyxl.load_workbook(excel_file)
        # getting all sheets
        sheets = wb.sheetnames
        print(sheets)
        # getting a particular sheet
        worksheet = wb["Sheet1"]
        print(worksheet)
        # getting active sheet
        active_sheet = wb.active
        print(active_sheet)
        # reading a cell
        print(worksheet["A1"].value)
        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
                print(cell.value)
            excel_data.append(row_data)
            Product_Details.objects.all().delete()
            Recommend_Prediction.objects.all().delete()
    for r in range(1, active_sheet.max_row+1):
        Product_Details.objects.create(

        idno= active_sheet.cell(r, 1).value,
        ProductId= active_sheet.cell(r, 2).value,
        UserId= active_sheet.cell(r, 3).value,
        ProfileName= active_sheet.cell(r, 4).value,
        HelpfulnessNumerator= active_sheet.cell(r, 5).value,
        HelpfulnessDenominator= active_sheet.cell(r, 6).value,
        Score= active_sheet.cell(r, 7).value,
        Time= active_sheet.cell(r, 8).value,
        Summary= active_sheet.cell(r, 9).value,
        Review= active_sheet.cell(r, 10).value )

    return render(request, 'RUser/Add_DataSet_Details.html', {"excel_data": excel_data})


def Register1(request):

    if request.method == "POST":
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        phoneno = request.POST.get('phoneno')
        country = request.POST.get('country')
        state = request.POST.get('state')
        city = request.POST.get('city')
        ClientRegister_Model.objects.create(username=username, email=email, password=password, phoneno=phoneno,
                                            country=country, state=state, city=city)

        return render(request, 'RUser/Register1.html')
    else:
        return render(request,'RUser/Register1.html')

def ViewYourProfile(request):
    userid = request.session['userid']
    obj = ClientRegister_Model.objects.get(id= userid)
    return render(request,'RUser/ViewYourProfile.html',{'object':obj})


def Search_DataSets(request):
    if request.method == "POST":
        kword = request.POST.get('keyword')
        if request.method == "POST":
            kword = request.POST.get('keyword')
            df = pd.read_csv('Reviews.csv')
            df
            df.columns
            df.rename(columns={'Score': 'Rating', 'Text': 'Review'}, inplace=True)

            def apply_recommend(Rating):
                if (Rating <= 2):
                    return 0  # No Recommend
                else:
                    return 1  # Recommend

            df['recommend'] = df['Rating'].apply(apply_recommend)
            df.drop(['Rating'], axis=1, inplace=True)
            recommend = df['recommend'].value_counts()
            df.drop(
                ['Id', 'ProductId', 'UserId', 'ProfileName', 'HelpfulnessNumerator', 'HelpfulnessDenominator', 'Time',
                 'Summary'], axis=1, inplace=True)

            cv = CountVectorizer()
            X = df['Review']
            y = df['recommend']
            X = cv.fit_transform(X)

            models = []
            from sklearn.model_selection import train_test_split
            X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.20)
            X_train.shape, X_test.shape, y_train.shape

            print("Naive Bayes")

            from sklearn.naive_bayes import MultinomialNB
            NB = MultinomialNB()
            models.append(('naive_bayes', NB))
            # SVM Model
            print("SVM")
            from sklearn import svm
            lin_clf = svm.LinearSVC()
            lin_clf.fit(X_train, y_train)
            models.append(('svm', lin_clf))

            print("Logistic Regression")

            from sklearn.linear_model import LogisticRegression
            reg = LogisticRegression(random_state=0, solver='lbfgs').fit(X_train, y_train)
            models.append(('logistic', reg))

            classifier = VotingClassifier(models)
            classifier.fit(X_train, y_train)
            y_pred = classifier.predict(X_test)

            review_data = [kword]
            vector1 = cv.transform(review_data).toarray()
            predict_text = classifier.predict(vector1)

            pred = str(predict_text).replace("[", "")
            pred1 = pred.replace("]", "")

            prediction = int(pred1)

            if prediction == 0:
              predict = 'No Recommend'
            else:
              predict = 'Recommend'

        return render(request, 'RUser/Search_DataSets.html',{'objs': predict})
    return render(request, 'RUser/Search_DataSets.html')



